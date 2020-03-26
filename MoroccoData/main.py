# -*- coding: utf-8 -*-
"""
Created on Wed Mar 25 19:42:29 2020

@author: USER
"""

import requests
import bs4 as bs
from coronaclasses import country
import openpyxl 

#import urllib.request
#import time
#import sys

def get_data_depends_on_country(CountryName):
    tbody=soup.find('tbody')
    rows=tbody.find_all('tr')
    print("looking for the row")
    for row in rows:
        lign=row.td.text
        if(lign==CountryName):
            return row
    
def get_object_depends_on_row(row):
    lst=[]
    infos=row.find_all('td')
    print("looking for data")
    for info in infos:
        value=info.text
        lst.append(value)
    obj=country()
    obj.setdata(lst)
    return obj

def add_row_to_excel(filename,obj):
    print("add to excell file")
    vk=openpyxl.load_workbook('./data/'+filename+'.xlsx');
    sh=vk['Feuil1']
    rows=sh.max_row
    collumn=sh.max_column
    lst=obj.class_to_array()
    for j in range(1,collumn+1):
        sh.cell(rows+1,j).value=lst[j-1]
    
    vk.save('./data/'+filename+'.xlsx')
    
print('start programe')

filePath='/data'
url="https://www.worldometers.info/coronavirus/"

raw=requests.get(url).text
soup=bs.BeautifulSoup(raw, 'lxml')
Countryname='Morocco'
row=get_data_depends_on_country(Countryname)
obj=get_object_depends_on_row(row)
obj.print()
add_row_to_excel(Countryname,obj)

#/*************************************************************

print('done')






'''
vk=openpyxl.load_workbook('./data/China.xlsx');
sh=vk['Sheet1']
#print(sh.cell(1,1).value)
rows=sh.max_row
collumn=sh.max_column
print(str(rows))
#sh.cell(2,2).value=2
for i in range(1,rows+1):
    for j in range(1,collumn+1):
        c=sh.cell(i,j)
        #print(c.value)
sh.cell(rows+1,5).value=4#to add to new ligne

vk.save('./data/China.xlsx')
'''






